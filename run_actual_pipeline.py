#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
实收实付数据处理核心脚本

流程:
1. 拆分实收实付数据 - 从原始数据提取指定月份的实收实付明细
2. 生成实收实付数据 - 按客户名称汇总
3. 生成凭证分录 - 根据实收实付数据生成凭证模板格式的分录

参考 run_accounts_pipeline.py 的结构和复用逻辑
"""

import sys
import os
import pandas as pd
import sqlite3
from openpyxl import Workbook, load_workbook

# 添加当前目录到路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# ============================================================================
# 凭证模板字段顺序 (73个字段，从 run_accounts_pipeline.py 复用)
# ============================================================================
VOUCHER_FIELDS = [
    '会计期间', '凭证类别字', '凭证类别排序号', '凭证编号', '行号', '制单日期',
    '附单据数', '制单人', '审核人', '记账人', '记账标志', '出纳人', '凭证标志',
    '凭证头自定义项1', '凭证头自定义项2', '摘要', '科目编码', '币种', '借方金额',
    '贷方金额', '外币借方金额', '外币贷方金额', '汇率', '数量借方', '数量贷方',
    '结算方式编码', '票号', '票号发生日期', '部门编码', '职员编码', '客户编码',
    '供应商编码', '项目编码', '项目大类编码', '业务员', '对方科目编码', '银行帐两清标志',
    '往来帐两清标志', '是否核销', '外部凭证帐套号', '外部凭证会计年度', '外部凭证系统名称',
    '外部凭证系统版本号', '外部凭证制单日期', '外部凭证会计期间', '外部凭证业务类型',
    '外部凭证业务号', '日期', '标志', '外部凭证单据号', '凭证是否可修改', '凭证分录是否可增删',
    '凭证合计金额是否保值', '分录数值是否可修改', '分录科目是否可修改', '分录受控科目可用状态',
    '分录往来项是否可修改', '分录部门是否可修改', '分录项目是否可修改', '分录往来项是否必输',
    '自定义字段1', '自定义字段2', '自定义字段3', '自定义字段4', '自定义字段5',
    '自定义字段6', '自定义字段7', '自定义字段8', '自定义字段9', '自定义字段10',
    '现金项目编号', '现金借方', '现金贷方'
]

# ============================================================================
# 共用函数 (从 run_accounts_pipeline.py 复用)
# ============================================================================

def get_customer_code(db_path: str, customer_name: str) -> tuple:
    """根据客户名称获取客户编码（支持中英文括号差异）"""
    # 规范化名称：统一括号类型后再匹配
    normalized = customer_name.replace('（', '(').replace('）', ')')
    conn = sqlite3.connect(db_path)
    cursor = conn.execute(
        '''SELECT 客户编号 FROM customers
           WHERE REPLACE(REPLACE(客户名称, '（', '('), '）', ')') = ?
              OR REPLACE(REPLACE(客户简称, '（', '('), '）', ')') = ?
              OR REPLACE(REPLACE(总公司全称, '（', '('), '）', ')') = ?''',
        (normalized, normalized, normalized)
    )
    result = cursor.fetchone()
    conn.close()
    if result:
        return str(result[0]), True
    return '', False


def create_entry(month, year, voucher_no, row_no, summary, subject_code, debit, credit,
                 customer_code, customer_name, counter_subject, voucher_date, invoice_count=1,
                 supplier_code='', supplier_name=''):
    """
    创建单条分录 (从 run_accounts_pipeline.py 复用)
    字段结构和固定值逻辑保持不变
    """
    entry = {field: '' for field in VOUCHER_FIELDS}

    entry['会计期间'] = month
    entry['凭证类别字'] = '记'
    entry['凭证类别排序号'] = 1
    entry['凭证编号'] = voucher_no
    entry['行号'] = row_no
    entry['制单日期'] = voucher_date
    entry['附单据数'] = invoice_count
    entry['制单人'] = '1'
    entry['审核人'] = '4'
    entry['记账人'] = '4'
    entry['记账标志'] = 1
    entry['摘要'] = summary
    entry['科目编码'] = subject_code
    entry['借方金额'] = debit
    entry['贷方金额'] = credit
    entry['客户编码'] = customer_code
    entry['供应商编码'] = supplier_code
    entry['对方科目编码'] = counter_subject
    entry['银行帐两清标志'] = 0
    entry['往来帐两清标志'] = 0
    entry['是否核销'] = 0
    entry['日期'] = voucher_date
    entry['凭证是否可修改'] = 0
    entry['凭证分录是否可增删'] = 0
    entry['凭证合计金额是否保值'] = 0
    entry['分录数值是否可修改'] = 0
    entry['分录科目是否可修改'] = 0
    entry['分录受控科目可用状态'] = 0
    entry['分录往来项是否可修改'] = 0
    entry['分录部门是否可修改'] = 0
    entry['分录项目是否可修改'] = 0
    entry['分录往来项是否必输'] = 0
    entry['自定义字段1'] = customer_name if customer_name else supplier_name

    return entry


def save_voucher_excel(entries: list, output_file: str):
    """
    保存凭证分录到Excel (从 run_accounts_pipeline.py 复用)
    """
    df_out = pd.DataFrame(entries)
    df_out = df_out[VOUCHER_FIELDS]

    # 确保客户编码保存为字符串格式（保留前导0）
    df_out['客户编码'] = df_out['客户编码'].astype(str)
    df_out['供应商编码'] = df_out['供应商编码'].astype(str)

    # 清空自定义字段1
    df_out['自定义字段1'] = ''

    # 设置外币和数量字段为0
    for field in ['外币借方金额', '外币贷方金额', '数量借方', '数量贷方']:
        if field in df_out.columns:
            df_out[field] = 0

    # 设置汇率为1
    if '汇率' in df_out.columns:
        df_out['汇率'] = 1

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(list(df_out.columns))

    # 需要设为空值的字段
    empty_fields = [
        '借方金额', '贷方金额',
        '银行帐两清标志', '往来帐两清标志', '是否核销',
        '凭证是否可修改', '凭证分录是否可增删', '凭证合计金额是否保值',
        '分录数值是否可修改', '分录科目是否可修改', '分录受控科目可用状态',
        '分录往来项是否可修改', '分录部门是否可修改', '分录项目是否可修改',
        '分录往来项是否必输'
    ]

    for _, row in df_out.iterrows():
        row_data = list(row)
        for field in empty_fields:
            if field in df_out.columns:
                idx = df_out.columns.get_loc(field)
                val = row_data[idx]
                if val == 0 or val == '0' or val == 0.0 or val == '':
                    row_data[idx] = None
        ws.append(row_data)

    # 设置金额格式（仅对非空值）
    debit_col = list(df_out.columns).index('借方金额') + 1
    credit_col = list(df_out.columns).index('贷方金额') + 1
    for row in range(2, len(df_out) + 2):
        for col in [debit_col, credit_col]:
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.number_format = '0.00'

    wb.save(output_file)


# ============================================================================
# 第一步: 拆分实收实付数据（银行流水 -> 按月按对方单位汇总）
# ============================================================================

def step1_split_actual(bank_file: str, month: int, output_file: str) -> bool:
    """
    第一步: 拆分实收实付数据（银行流水）

    从银行流水文件读取数据，按指定月份过滤，按对方单位汇总转入/转出金额。

    Args:
        bank_file: 银行流水文件（如 4月预付.xlsx）
        month: 要处理的月份
        output_file: 输出文件路径

    Returns:
        是否成功
    """
    print("\n" + "=" * 70)
    print("第一步: 拆分实收实付数据（银行流水）")
    print("=" * 70)

    try:
        # 读取银行流水，假设第一行是表头
        df = pd.read_excel(bank_file, header=0)

        # 确保金额列为数值类型
        df['转入金额'] = pd.to_numeric(df['转入金额'], errors='coerce').fillna(0)
        df['转出金额'] = pd.to_numeric(df['转出金额'], errors='coerce').fillna(0)

        # 转换交易时间并提取月份
        df['交易时间'] = pd.to_datetime(df['交易时间'], errors='coerce')
        df['数据月份'] = df['交易时间'].dt.month

        # 过滤指定月份的数据
        month_df = df[df['数据月份'] == month].copy()

        # 按对方单位汇总（排除空值，即银行手续费和工资等）
        # NaN 或空白字符串都视为"其他支出"
        supplier_mask = month_df['对方单位'].notna() & (month_df['对方单位'].astype(str).str.strip() != '')
        supplier_df = month_df[supplier_mask].copy()

        # 银行手续费和工资等（对方单位为空/NaN）单独统计
        other_df = month_df[~supplier_mask].copy()
        other_summary = other_df.groupby('摘要')['转出金额'].agg(['sum', 'count']).reset_index()
        other_summary.columns = ['摘要', '金额', '笔数']
        other_summary = other_summary.sort_values('金额', ascending=False)

        grouped = supplier_df.groupby('对方单位').agg(
            转入笔数=('转入金额', lambda x: (x > 0).sum()),
            转出笔数=('转出金额', lambda x: (x > 0).sum()),
            转入总额=('转入金额', 'sum'),
            转出总额=('转出金额', 'sum')
        ).reset_index()

        # 计算净额
        grouped['净额'] = grouped['转入总额'] - grouped['转出总额']
        grouped['月份'] = month

        # 拆分为转入和转出两个DataFrame
        # 转入：对方单位 + 转入笔数 + 转入总额 + 净额
        df_in = grouped[grouped['转入笔数'] > 0][['对方单位', '月份', '转入笔数', '转入总额', '净额']].copy()
        df_in = df_in.sort_values('转入总额', ascending=False).reset_index(drop=True)

        # 转出：对方单位 + 转出笔数 + 转出总额 + 净额
        df_out = grouped[grouped['转出笔数'] > 0][['对方单位', '月份', '转出笔数', '转出总额', '净额']].copy()
        df_out = df_out.sort_values('转出总额', ascending=False).reset_index(drop=True)

        # 保存到Excel，三个sheet
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_in.to_excel(writer, sheet_name='转入汇总', index=False)
            df_out.to_excel(writer, sheet_name='转出汇总', index=False)
            other_summary.to_excel(writer, sheet_name='其他支出', index=False)

        print(f"  转入汇总: {len(df_in)} 个单位, 总额 {df_in['转入总额'].sum():,.2f}")
        print(f"  转出汇总: {len(df_out)} 个单位, 总额 {df_out['转出总额'].sum():,.2f}")
        if len(other_summary) > 0:
            print(f"  其他支出: {len(other_summary)} 类, 总额 {other_summary['金额'].sum():,.2f}")

        print(f"\n已保存到: {output_file}")
        return True

    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()
        return False


# ============================================================================
# 第二步: 生成实收实付数据 (TODO: 根据业务需求实现)
# ============================================================================

def step2_generate_actual(input_file: str, month: int, output_file: str) -> bool:
    """
    第二步: 生成实收实付数据

    TODO: 根据实收实付的业务需求实现

    Args:
        input_file: 拆分后的数据文件
        month: 月份
        output_file: 输出文件路径

    Returns:
        是否成功
    """
    print("\n" + "=" * 70)
    print("第二步: 生成实收实付数据")
    print("=" * 70)
    print("TODO: 实现实收实付数据生成逻辑")
    return False


# ============================================================================
# 第三步: 生成凭证分录 (TODO: 根据实收实付业务逻辑实现)
# ============================================================================

def step3_generate_actual_voucher(detail_file: str, db_path: str, month: int, year: int,
                                   voucher_no: int, output_file: str, unmatched_file: str) -> tuple:
    """
    第三步: 生成实收实付凭证分录（转入 = 收回应收账款）

    借方: 银行存款(102) 汇总为1行
    贷方: 应收账款(122) 每单位1行

    Args:
        detail_file: 实收实付明细文件（step1输出，含有转入汇总sheet）
        db_path: 数据库路径
        month: 月份
        year: 年份
        voucher_no: 凭证编号
        output_file: 输出文件路径
        unmatched_file: 未匹配客户导出路径

    Returns:
        (是否成功, 未匹配客户列表)
    """
    print("\n" + "=" * 70)
    print("第三步: 生成实收实付凭证分录（转入）")
    print("=" * 70)

    try:
        from datetime import datetime
        voucher_date = datetime.now().strftime('%Y-%m-%d')

        # 读取转入汇总 sheet
        df = pd.read_excel(detail_file, sheet_name='转入汇总')
        print(f"读取转入汇总: {len(df)} 个单位")

        entries = []
        row_no = 1
        unmatched_customers = []
        total_in_amount = 0.0
        total_invoice_count = 0

        # 贷方分录: 每单位1行（应收账款）
        for _, row in df.iterrows():
            unit_name = str(row['对方单位'])
            in_amount = float(row['转入总额'])
            in_count = int(row['转入笔数'])

            customer_code, matched = get_customer_code(db_path, unit_name)

            if not matched:
                unmatched_customers.append({
                    '对方单位': unit_name,
                    '转入总额': in_amount,
                    '转入笔数': in_count
                })

            # 累计汇总金额
            total_in_amount += in_amount
            total_invoice_count += in_count

            # 贷方: 应收账款 (122)
            entry = create_entry(
                month=month, year=year, voucher_no=voucher_no, row_no=row_no,
                summary=f'{month}月收回应收账款', subject_code='122',
                debit=0, credit=in_amount,
                customer_code=customer_code, customer_name=unit_name,
                counter_subject='102', voucher_date=voucher_date,
                invoice_count=in_count
            )
            entries.append(entry)
            row_no += 1

        # 借方分录: 银行存款汇总成1行（放在贷方分录之后）
        entry = create_entry(
            month=month, year=year, voucher_no=voucher_no, row_no=row_no,
            summary=f'{month}月收回应收账款', subject_code='102',
            debit=total_in_amount, credit=0,
            customer_code='', customer_name='',
            counter_subject='122', voucher_date=voucher_date,
            invoice_count=total_invoice_count
        )
        entries.append(entry)

        # 计算合计验证
        debit_total = sum(float(e['借方金额']) if e['借方金额'] else 0 for e in entries)
        credit_total = sum(float(e['贷方金额']) if e['贷方金额'] else 0 for e in entries)

        print(f"生成凭证分录: {len(df)} 条贷方 + 1 条借方 = {len(entries)} 条")
        print(f"  借方合计(102银行存款): {debit_total:,.2f}")
        print(f"  贷方合计(122应收账款): {credit_total:,.2f}")

        # 保存凭证
        save_voucher_excel(entries, output_file)
        print(f"\n已保存到: {output_file}")

        # 保存未匹配客户
        if unmatched_customers:
            df_unmatched = pd.DataFrame(unmatched_customers)
            df_unmatched = df_unmatched.sort_values('转入总额', ascending=False)
            df_unmatched.to_excel(unmatched_file, index=False, sheet_name='未匹配客户')
            print(f"未匹配客户已导出到: {unmatched_file}")

        return True, unmatched_customers

    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()
        return False, []


# ============================================================================
# 第三步B: 生成付款凭证分录（转出 = 预付账款）
# ============================================================================

def step3_generate_payment_voucher(detail_file: str, db_path: str, month: int, year: int,
                                     voucher_no: int, output_file: str, unmatched_file: str) -> tuple:
    """
    第三步B: 生成付款凭证分录（转出 = 预付账款）

    借方: 预付账款(126) 每单位1行
    贷方: 银行存款(102) 汇总1行

    Args:
        detail_file: 实收实付明细文件（step1输出，含有转出汇总sheet）
        db_path: 数据库路径
        month: 月份
        year: 年份
        voucher_no: 凭证编号
        output_file: 输出文件路径
        unmatched_file: 未匹配客户导出路径

    Returns:
        (是否成功, 未匹配客户列表)
    """
    print("\n" + "=" * 70)
    print("第三步: 生成付款凭证分录（转出）")
    print("=" * 70)

    try:
        from datetime import datetime
        voucher_date = datetime.now().strftime('%Y-%m-%d')

        # 读取转出汇总 sheet
        df = pd.read_excel(detail_file, sheet_name='转出汇总')
        print(f"读取转出汇总: {len(df)} 个单位")

        entries = []
        row_no = 1
        unmatched_customers = []
        total_out_amount = 0.0
        total_invoice_count = 0

        # 借方分录: 每单位1行（预付账款）
        for _, row in df.iterrows():
            unit_name = str(row['对方单位'])
            out_amount = float(row['转出总额'])
            out_count = int(row['转出笔数'])

            customer_code, matched = get_customer_code(db_path, unit_name)

            if not matched:
                unmatched_customers.append({
                    '对方单位': unit_name,
                    '转出总额': out_amount,
                    '转出笔数': out_count
                })

            # 累计汇总金额
            total_out_amount += out_amount
            total_invoice_count += out_count

            # 借方: 预付账款 (126)
            entry = create_entry(
                month=month, year=year, voucher_no=voucher_no, row_no=row_no,
                summary=f'{month}月预付账款', subject_code='126',
                debit=out_amount, credit=0,
                customer_code=customer_code, customer_name=unit_name,
                supplier_code='', supplier_name='',
                counter_subject='102', voucher_date=voucher_date,
                invoice_count=out_count
            )
            entries.append(entry)
            row_no += 1

        # 贷方分录: 银行存款汇总成1行（放在借方分录之后）
        entry = create_entry(
            month=month, year=year, voucher_no=voucher_no, row_no=row_no,
            summary=f'{month}月预付账款', subject_code='102',
            debit=0, credit=total_out_amount,
            customer_code='', customer_name='',
            counter_subject='126', voucher_date=voucher_date,
            invoice_count=total_invoice_count
        )
        entries.append(entry)

        # 计算合计验证
        debit_total = sum(float(e['借方金额']) if e['借方金额'] else 0 for e in entries)
        credit_total = sum(float(e['贷方金额']) if e['贷方金额'] else 0 for e in entries)

        print(f"生成凭证分录: {len(df)} 条借方 + 1 条贷方 = {len(entries)} 条")
        print(f"  借方合计(126预付账款): {debit_total:,.2f}")
        print(f"  贷方合计(102银行存款): {credit_total:,.2f}")

        # 保存凭证
        save_voucher_excel(entries, output_file)
        print(f"\n已保存到: {output_file}")

        # 保存未匹配客户
        if unmatched_customers:
            df_unmatched = pd.DataFrame(unmatched_customers)
            df_unmatched = df_unmatched.sort_values('转出总额', ascending=False)
            df_unmatched.to_excel(unmatched_file, index=False, sheet_name='未匹配客户')
            print(f"未匹配客户已导出到: {unmatched_file}")

        return True, unmatched_customers

    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()
        return False, []


# ============================================================================
# 主流程
# ============================================================================

def run_actual_pipeline(target_month: int, year: int = 2026, voucher_no: int = 1,
                       invoice_file: str = 'Docs/实收实付数据.xlsx',
                       db_path: str = 'data.db',
                       output_dir: str = 'Docs'):
    """
    执行实收实付全流程

    TODO: 实现完整流程

    Args:
        target_month: 要处理的月份
        year: 年份
        voucher_no: 凭证编号
        invoice_file: 原始实收实付数据文件
        db_path: 数据库路径
        output_dir: 输出目录
    """
    print("\n" + "=" * 70)
    print("实收实付数据处理全流程")
    print("=" * 70)
    print(f"处理月份: {target_month}月")
    print(f"年份: {year}")
    print(f"数据文件: {invoice_file}")
    print(f"输出目录: {output_dir}")
    print()
    print("TODO: 实收实付流程尚未实现")


def main():
    import argparse

    parser = argparse.ArgumentParser(description='实收实付数据处理全流程')
    parser.add_argument('target_month', type=int, help='要处理的月份 (1-12)')
    parser.add_argument('--year', type=int, default=2026, help='年份 (默认2026)')
    parser.add_argument('--voucher-no', type=int, default=1, help='凭证编号 (默认1)')
    parser.add_argument('--invoice-file', default='Docs/实收实付数据.xlsx', help='原始数据文件')
    parser.add_argument('--db', default='data.db', help='SQLite数据库路径')
    parser.add_argument('--output-dir', default='Docs', help='输出目录 (默认Docs)')

    args = parser.parse_args()

    if args.target_month < 1 or args.target_month > 12:
        print("错误: 月份必须在1-12之间")
        sys.exit(1)

    run_actual_pipeline(
        target_month=args.target_month,
        year=args.year,
        voucher_no=args.voucher_no,
        invoice_file=args.invoice_file,
        db_path=args.db,
        output_dir=args.output_dir
    )


if __name__ == '__main__':
    main()