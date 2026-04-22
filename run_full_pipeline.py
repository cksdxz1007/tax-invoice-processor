#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
发票处理全流程自动化脚本

流程:
1. 拆分发票数据 - 从原始发票文件提取指定月份的发票明细
2. 生成应收数据 - 合并专票普票，按客户名称汇总
3. 生成凭证分录 - 根据应收数据生成凭证模板格式的分录

使用方式:
    python run_full_pipeline.py <目标月份> [选项]

示例:
    python run_full_pipeline.py 4    # 处理4月发票数据
    python run_full_pipeline.py 4 --check  # 处理4月数据并检查未匹配客户
"""

import sys
import os
import pandas as pd
import sqlite3
from openpyxl import Workbook, load_workbook

# 添加当前目录到路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# ============================================================================
# 颜色常量
# ============================================================================
YELLOW_COLORS = {'FFFF00', 'FFFFFF00', 'FF00FFFF00'}  # 黄色填充


def is_yellow_fill(cell) -> bool:
    """检查单元格是否为黄色填充"""
    try:
        if cell.fill and cell.fill.fgColor:
            color = cell.fill.fgColor.rgb
            if color:
                # 转换为字符串处理
                color = str(color).upper()
                return 'FFFF00' in color or color in YELLOW_COLORS
    except:
        pass
    return False


# ============================================================================
# 第一步: 拆分发票数据
# ============================================================================
SECTION_MARKERS = {
    '专票': '本月开出专票',
    '普票': '本月开出普票',
    '进项': '本月收到进项',
}
SUMMARY_KEYWORDS = ['小计', '合计', '统计', '汇总']
VOUCHER_FIELDS = [
    '会计期间', '凭证类别字', '凭证类别排序号', '凭证编号', '行号', '制单日期',
    '附单据数', '制单人', '审核人', '记账人', '记账标志', '出纳人', '凭证标志',
    '凭证头自定义项1', '凭证头自定义项2', '摘要', '科目编码', '币种', '借方金额',
    '贷方金额', '外币借方金额', '外币贷方金额', '汇率', '数量借方', '数量贷方',
    '结算方式编码', '票号', '票号发生日期', '部门编码', '职员编码', '客户编码',
    '供应商编码', '项目编码', '项目大类编码', '业务员', '对方科目编码', '银行帐两清标志',
    '往来帐两清标志', '是否核销', '外部凭证帐套号', '外部凭证会计年度', '外部凭证系统名称',
    '外部凭证系统版本号', '外部凭证制单日期', '外部凭证会计期间', '外部凭证业务类型',
    '外部凭证业务号', '日期', '标志', '外部凭证单据号', '凭证是否可修改', '凭证分录是否可增删',
    '凭证合计金额是否保值', '分录数值是否可修改', '分录科目是否可修改', '分录受控科目可用状态',
    '分录往来项是否可修改', '分录部门是否可修改', '分录项目是否可修改', '分录往来项是否必输',
    '自定义字段1', '自定义字段2', '自定义字段3', '自定义字段4', '自定义字段5',
    '自定义字段6', '自定义字段7', '自定义字段8', '自定义字段9', '自定义字段10',
    '现金项目编号', '现金借方', '现金贷方'
]


def find_section_boundaries(df):
    """动态识别三个分区的起止位置"""
    boundaries = {}
    first_col = df.iloc[:, 0].astype(str)

    for section, marker in SECTION_MARKERS.items():
        matches = df[first_col.str.contains(marker, na=False)].index.tolist()
        if matches:
            boundaries[section] = {'start': matches[0]}

    sections = list(boundaries.keys())
    for i, section in enumerate(sections):
        if i < len(sections) - 1:
            next_section = sections[i + 1]
            boundaries[section]['end'] = boundaries[next_section]['start'] - 1
        else:
            last_data_row = boundaries[section]['start']
            for idx in range(boundaries[section]['start'] + 1, len(df)):
                row_val = df.iloc[idx, 0]
                if pd.isna(row_val) or any(kw in str(row_val) for kw in SUMMARY_KEYWORDS):
                    last_data_row = idx - 1
                    break
                last_data_row = idx
            boundaries[section]['end'] = last_data_row

    return boundaries


def find_column_row(df, start, end):
    """在指定范围内查找列名行"""
    for idx in range(start, min(end + 1, start + 5)):
        row_values = df.iloc[idx].astype(str).tolist()
        if '单位名称' in row_values and '日期' in row_values:
            return idx
    return start


def step1_split_invoice(invoice_file: str, month: int, output_file: str) -> bool:
    """
    第一步: 拆分发票数据

    Args:
        invoice_file: 原始发票Excel文件
        month: 要处理的月份
        output_file: 输出文件路径

    Returns:
        是否成功
    """
    print("\n" + "=" * 70)
    print("第一步: 拆分发票数据")
    print("=" * 70)

    try:
        # 使用 openpyxl 读取以支持颜色检测
        wb = load_workbook(invoice_file, data_only=True)
        ws = wb[f'{month}月']

        # 获取原始数据用于边界识别
        df = pd.read_excel(invoice_file, sheet_name=f'{month}月', header=None)
        print(f"读取工作表 '{month}月': {df.shape[0]} 行 x {df.shape[1]} 列")

        boundaries = find_section_boundaries(df)
        print(f"识别分区: {list(boundaries.keys())}")

        # 找到「金额」列的索引
        columns = ['单位名称', '日期', '发票张数', '发票号', '单票合计', '金额', '税额', '备注']
        amount_col_idx = None
        for i, col in enumerate(columns):
            if col == '金额':
                amount_col_idx = i
                break

        sheets = {}

        for section, boundary in boundaries.items():
            col_row = find_column_row(df, boundary['start'], boundary['end'])
            boundary['col_row'] = col_row
            data_start = col_row + 1

            section_df = df.iloc[data_start:boundary['end'] + 1].copy()
            section_df.columns = df.iloc[col_row].tolist()
            section_df = section_df.reset_index(drop=True)

            # 过滤汇总行
            section_df = section_df[
                section_df['单位名称'].notna() &
                ~section_df['单位名称'].astype(str).str.contains('|'.join(SUMMARY_KEYWORDS), na=False)
            ]

            # 过滤黄色填充的行（金额列为黄色）
            if amount_col_idx is not None:
                original_rows = []
                yellow_count = 0
                for idx, _ in section_df.iterrows():
                    # 计算原始DataFrame中的位置对应的Excel行号
                    # df的索引从0开始，Excel从1开始
                    excel_row = data_start + idx + 1
                    excel_col = amount_col_idx + 1
                    cell = ws.cell(row=excel_row, column=excel_col)
                    if is_yellow_fill(cell):
                        yellow_count += 1
                    else:
                        original_rows.append(idx)

                section_df = section_df.loc[original_rows]
                if yellow_count > 0:
                    print(f"  {section}: 过滤 {yellow_count} 行黄色填充数据")

            # 添加汇总行
            summary_row = {
                '单位名称': '合计',
                '日期': '',
                '发票张数': section_df['发票张数'].sum() if '发票张数' in section_df.columns else 0,
                '发票号': '',
                '单票合计': section_df['单票合计'].sum() if '单票合计' in section_df.columns else 0,
                '金额': section_df['金额'].sum() if '金额' in section_df.columns else 0,
                '税额': section_df['税额'].sum() if '税额' in section_df.columns else 0,
                '备注': ''
            }
            section_df = pd.concat([section_df, pd.DataFrame([summary_row])], ignore_index=True)

            sheets[section] = section_df
            print(f"  {section}: {len(section_df) - 1} 条数据 + 1 行汇总")

        # 保存
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for section, df_section in sheets.items():
                df_section.to_excel(writer, sheet_name=section, index=False)

        print(f"\n已保存到: {output_file}")
        return True

    except Exception as e:
        print(f"错误: {e}")
        return False


# ============================================================================
# 第二步: 生成应收/应付数据
# ============================================================================
def step2_generate_receivable(input_file: str, month: int, output_file: str) -> bool:
    """
    第二步: 生成应收数据和应付数据

    Args:
        input_file: 发票明细文件
        month: 月份
        output_file: 输出文件路径

    Returns:
        是否成功
    """
    print("\n" + "=" * 70)
    print("第二步: 生成应收/应付数据")
    print("=" * 70)

    try:
        # 读取专票和普票 -> 应收数据
        df_vat = pd.read_excel(input_file, sheet_name='专票')
        df_vat = df_vat[df_vat['单位名称'] != '合计']
        df_normal = pd.read_excel(input_file, sheet_name='普票')
        df_normal = df_normal[df_normal['单位名称'] != '合计']

        print(f"专票: {len(df_vat)} 条")
        print(f"普票: {len(df_normal)} 条")

        df_receivable = pd.concat([df_vat, df_normal], ignore_index=True)
        print(f"合并后(应收): {len(df_receivable)} 条")

        receivable = df_receivable.groupby('单位名称').agg({
            '单票合计': 'sum',
            '金额': 'sum',
            '税额': 'sum',
            '发票张数': 'sum'
        }).reset_index()

        receivable.columns = ['单位名称', '单票合计', '金额', '税额', '发票张数']
        receivable = receivable.sort_values('单票合计', ascending=False).reset_index(drop=True)

        print(f"\n应收数据汇总: {len(receivable)} 个客户")
        print(f"  总金额: {receivable['金额'].sum():,.2f}")
        print(f"  总税额: {receivable['税额'].sum():,.2f}")

        # 读取进项 -> 应付数据
        if '进项' in pd.ExcelFile(input_file).sheet_names:
            df_input = pd.read_excel(input_file, sheet_name='进项')
            df_input = df_input[df_input['单位名称'] != '合计']
            print(f"\n进项: {len(df_input)} 条")

            payable = df_input.groupby('单位名称').agg({
                '单票合计': 'sum',
                '金额': 'sum',
                '税额': 'sum',
                '发票张数': 'sum'
            }).reset_index()

            payable.columns = ['单位名称', '单票合计', '金额', '税额', '发票张数']
            payable = payable.sort_values('单票合计', ascending=False).reset_index(drop=True)

            print(f"应付数据汇总: {len(payable)} 个供应商")
            print(f"  总金额: {payable['金额'].sum():,.2f}")
            print(f"  总税额: {payable['税额'].sum():,.2f}")
        else:
            payable = pd.DataFrame(columns=['单位名称', '单票合计', '金额', '税额', '发票张数'])
            print(f"\n未找到进项数据")

        # 保存到Excel
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            receivable.to_excel(writer, sheet_name='应收数据', index=False)
            payable.to_excel(writer, sheet_name='应付数据', index=False)

        print(f"\n已保存到: {output_file}")
        print(f"  - 【应收数据】sheet: {len(receivable)} 个客户")
        print(f"  - 【应付数据】sheet: {len(payable)} 个供应商")
        return True

    except Exception as e:
        print(f"错误: {e}")
        return False


# ============================================================================
# 第三步: 生成凭证分录
# ============================================================================
def get_customer_code(db_path: str, customer_name: str) -> tuple:
    """根据客户名称获取客户编码"""
    conn = sqlite3.connect(db_path)
    cursor = conn.execute(
        'SELECT 客户编号 FROM customers WHERE 客户名称 = ? OR 客户简称 = ? OR 总公司全称 = ?',
        (customer_name, customer_name, customer_name)
    )
    result = cursor.fetchone()
    conn.close()
    if result:
        return str(result[0]), True
    return '', False


def create_entry(month, year, voucher_no, row_no, summary, subject_code, debit, credit,
                 customer_code, customer_name, counter_subject, voucher_date, invoice_count=1):
    """创建单条分录"""
    entry = {field: '' for field in VOUCHER_FIELDS}

    entry['会计期间'] = month
    entry['凭证类别字'] = '记'
    entry['凭证类别排序号'] = 1
    entry['凭证编号'] = voucher_no
    entry['行号'] = row_no
    entry['制单日期'] = voucher_date
    entry['附单据数'] = invoice_count
    entry['制单人'] = '1'
    entry['审核人'] = '4'
    entry['记账人'] = '4'
    entry['记账标志'] = 1
    entry['摘要'] = summary
    entry['科目编码'] = subject_code
    entry['借方金额'] = debit
    entry['贷方金额'] = credit
    entry['客户编码'] = customer_code
    entry['对方科目编码'] = counter_subject
    entry['银行帐两清标志'] = 0
    entry['往来帐两清标志'] = 0
    entry['是否核销'] = 0
    entry['日期'] = voucher_date
    entry['凭证是否可修改'] = 0
    entry['凭证分录是否可增删'] = 0
    entry['凭证合计金额是否保值'] = 0
    entry['分录数值是否可修改'] = 0
    entry['分录科目是否可修改'] = 0
    entry['分录受控科目可用状态'] = 0
    entry['分录往来项是否可修改'] = 0
    entry['分录部门是否可修改'] = 0
    entry['分录项目是否可修改'] = 0
    entry['分录往来项是否必输'] = 0
    entry['自定义字段1'] = customer_name

    return entry


def step3_generate_voucher(invoice_file: str, db_path: str, month: int, year: int,
                            voucher_no: int, output_file: str, unmatched_file: str) -> tuple:
    """
    第三步: 生成凭证分录

    每行应收数据拆分为3条分录:
    1. 借方 应收账款 (单票合计)
    2. 贷方 商品销售收入 (金额)
    3. 贷方 销项税 (税额)

    Args:
        invoice_file: 发票明细文件
        db_path: 数据库路径
        month: 月份
        year: 年份
        voucher_no: 凭证编号
        output_file: 输出文件路径
        unmatched_file: 未匹配客户导出路径

    Returns:
        (是否成功, 未匹配客户列表)
    """
    print("\n" + "=" * 70)
    print("第三步: 生成凭证分录")
    print("=" * 70)

    try:
        voucher_date = f'{year}-{month:02d}-01'
        df = pd.read_excel(invoice_file, sheet_name='应收数据')
        # 排除汇总行
        df = df[df['单位名称'] != '合计']
        print(f"读取应收数据: {len(df)} 个客户")

        entries = []
        row_no = 1
        unmatched_customers = []

        # 每行应收数据拆分为3条分录
        for _, row in df.iterrows():
            customer_name = str(row['单位名称'])
            customer_code, matched = get_customer_code(db_path, customer_name)
            single_total = float(row['单票合计'])  # 价税合计
            amount = float(row['金额'])  # 不含税金额
            tax = float(row['税额'])  # 税额
            invoice_count = int(row['发票张数'])  # 发票张数

            if not matched:
                unmatched_customers.append({
                    '单位名称': customer_name,
                    '单票合计': single_total,
                    '金额': amount,
                    '税额': tax
                })

            # 分录1: 借方 应收账款 (单票合计)
            entry = create_entry(
                month=month, year=year, voucher_no=voucher_no, row_no=row_no,
                summary=f'{month}月应收账款', subject_code='122',
                debit=single_total, credit=0,
                customer_code=customer_code, customer_name=customer_name,
                counter_subject='501,2210102', voucher_date=voucher_date,
                invoice_count=invoice_count
            )
            entries.append(entry)
            row_no += 1

            # 分录2: 贷方 商品销售收入 (金额)
            entry = create_entry(
                month=month, year=year, voucher_no=voucher_no, row_no=row_no,
                summary=f'{month}月销售收入', subject_code='501',
                debit=0, credit=amount,
                customer_code=customer_code, customer_name=customer_name,
                counter_subject='122', voucher_date=voucher_date,
                invoice_count=invoice_count
            )
            entries.append(entry)
            row_no += 1

            # 分录3: 贷方 销项税 (税额)
            entry = create_entry(
                month=month, year=year, voucher_no=voucher_no, row_no=row_no,
                summary=f'{month}月销项税', subject_code='2210102',
                debit=0, credit=tax,
                customer_code=customer_code, customer_name=customer_name,
                counter_subject='122', voucher_date=voucher_date,
                invoice_count=invoice_count
            )
            entries.append(entry)
            row_no += 1

        # 计算合计
        debit_total = sum(float(e['借方金额']) if e['借方金额'] else 0 for e in entries)
        credit_total = sum(float(e['贷方金额']) if e['贷方金额'] else 0 for e in entries)

        print(f"生成凭证分录: {len(entries)} 条 (每客户3条)")
        print(f"  借方合计: {debit_total:,.2f}")
        print(f"  贷方合计: {credit_total:,.2f}")

        # 保存Excel
        df_out = pd.DataFrame(entries)
        df_out = df_out[VOUCHER_FIELDS]

        # 确保客户编码保存为字符串格式（保留前导0）
        df_out['客户编码'] = df_out['客户编码'].astype(str)
        df_out['供应商编码'] = df_out['供应商编码'].astype(str)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        ws.append(list(df_out.columns))

        for _, row in df_out.iterrows():
            ws.append(list(row))

        # 设置金额格式
        debit_col = list(df_out.columns).index('借方金额') + 1
        credit_col = list(df_out.columns).index('贷方金额') + 1
        for row in range(2, len(df_out) + 2):
            for col in [debit_col, credit_col]:
                ws.cell(row=row, column=col).number_format = '0.00'

        wb.save(output_file)

        print(f"\n已保存到: {output_file}")
        print(f"列数: {len(df_out.columns)} (模板要求: 73)")
        print(f"列名与模板一致: {list(df_out.columns) == VOUCHER_FIELDS}")

        # 保存未匹配客户
        if unmatched_customers:
            df_unmatched = pd.DataFrame(unmatched_customers)
            df_unmatched = df_unmatched.sort_values('单票合计', ascending=False)
            df_unmatched.to_excel(unmatched_file, index=False, sheet_name='未匹配客户')
            print(f"未匹配客户已导出到: {unmatched_file}")

        return True, unmatched_customers

    except Exception as e:
        print(f"错误: {e}")
        return False, []


# ============================================================================
# 主流程
# ============================================================================
def run_full_pipeline(target_month: int, year: int = 2026, voucher_no: int = 1,
                      invoice_file: str = 'Docs/2026年发票统计.xlsx',
                      db_path: str = 'data.db',
                      output_dir: str = 'Docs'):
    """
    执行全流程

    Args:
        target_month: 要处理的月份
        year: 年份
        voucher_no: 凭证编号
        invoice_file: 原始发票文件
        db_path: 数据库路径
        output_dir: 输出目录
    """
    # 检查发票文件是否存在
    if not os.path.exists(invoice_file):
        print(f"错误: 发票文件不存在: {invoice_file}")
        return False

    # 检查月份工作表是否存在
    try:
        sheets = pd.ExcelFile(invoice_file).sheet_names
        sheet_name = f'{target_month}月'
        if sheet_name not in sheets:
            print(f"错误: 文件中未找到工作表 '{sheet_name}'")
            print(f"可用工作表: {sheets}")
            return False
    except Exception as e:
        print(f"错误: 无法读取发票文件: {e}")
        return False

    # 确保输出目录存在
    os.makedirs(output_dir, exist_ok=True)

    print("=" * 70)
    print("发票处理全流程自动化")
    print("=" * 70)
    print(f"处理月份: {target_month}月")
    print(f"年份: {year}")
    print(f"发票文件: {invoice_file}")
    print(f"输出目录: {output_dir}")
    print()

    # 输出文件
    invoice_detail_file = f'{output_dir}/{target_month}月发票明细.xlsx'
    voucher_file = f'{output_dir}/{target_month}月应收凭证分录.xlsx'
    unmatched_file = f'{output_dir}/{target_month}月未匹配客户.xlsx'

    # 执行步骤
    results = {}

    # 第一步
    results['step1'] = step1_split_invoice(invoice_file, target_month, invoice_detail_file)

    # 第二步
    if results['step1']:
        results['step2'] = step2_generate_receivable(invoice_detail_file, target_month, invoice_detail_file)
    else:
        results['step2'] = False

    # 第三步
    if results['step1'] and results['step2']:
        results['step3'], unmatched = step3_generate_voucher(
            invoice_detail_file, db_path, target_month, year, voucher_no,
            voucher_file, unmatched_file
        )
    else:
        results['step3'] = False
        unmatched = []

    # 结果汇总
    print("\n" + "=" * 70)
    print("执行结果汇总")
    print("=" * 70)
    print(f"第一步 拆分发票数据: {'成功' if results.get('step1') else '失败'} ✓")
    print(f"第二步 生成应收数据: {'成功' if results.get('step2') else '失败'} ✓")
    print(f"第三步 生成凭证分录: {'成功' if results.get('step3') else '失败'} ✓")

    print("\n生成文件:")
    print(f"  发票明细: {invoice_detail_file}")
    print(f"  凭证分录: {voucher_file}")

    if unmatched:
        print(f"\n⚠ 警告: 有 {len(unmatched)} 个客户在客户档案中未找到匹配")
        print(f"  详情见: {unmatched_file}")

    print("\n流程完成!")


def main():
    import argparse

    parser = argparse.ArgumentParser(description='发票处理全流程自动化')
    parser.add_argument('target_month', type=int, help='要处理的月份 (1-12)')
    parser.add_argument('--year', type=int, default=2026, help='年份 (默认2026)')
    parser.add_argument('--voucher-no', type=int, default=1, help='凭证编号 (默认1)')
    parser.add_argument('--invoice-file', default='Docs/2026年发票统计.xlsx', help='原始发票文件')
    parser.add_argument('--db', default='data.db', help='SQLite数据库路径')
    parser.add_argument('--output-dir', default='Docs', help='输出目录 (默认Docs)')

    args = parser.parse_args()

    if args.target_month < 1 or args.target_month > 12:
        print("错误: 月份必须在1-12之间")
        sys.exit(1)

    run_full_pipeline(
        target_month=args.target_month,
        year=args.year,
        voucher_no=args.voucher_no,
        invoice_file=args.invoice_file,
        db_path=args.db,
        output_dir=args.output_dir
    )


if __name__ == '__main__':
    main()
